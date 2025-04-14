import tkinter as tk
from tkinter import font
from tkinter import ttk
from tkinter import messagebox
import os
import time
from datetime import datetime, timedelta, timezone
import threading
import xml.etree.ElementTree as ET
import re
from bs4 import BeautifulSoup
import pandas as pd
import traceback
import glob
import shutil
from dateutil.relativedelta import relativedelta
from pathlib import Path
import locale
from openpyxl import load_workbook
from dateutil.parser import parse

# Configuración regional para fechas en español
locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')

# Patrones de regex para identificar archivos/carpetas de muestras, calibraciones, etc.
patterns = [
    r'(\d{2}-\d{2}-\d{2})_SAMPLE (\d+)(?:\.D)?', # Ej: "25-06-24_SAMPLE 123.D"
    r'(\d{2}-\d{2}-\d{4})_CAL (\d+)(?:\.D)?',
    r'(\d{2}-\d{2}-\d{4})_SAMPLE (\d+)(?:\.D)?',
    r'(\d{2}-\d{2}-\d{4})_LIMP (\d+)(?:\.D)?',
    r'(\d{2}-\d{2}-\d{4})_QC (\d+)(?:\.D)?',
    r'(\d{2}-\d{2}-\d{4})_DOPA (\d+)(?:\.D)?',
    r'(\d{2}-\d{2}-\d{2})_TBE (\d+)(?:\.D)?',
    r'^\d{2}-\d{2}-\d{2}_TPH.*\.sirslt$',
    r'^(Set_CG-\d{3}-a_(Front|Back)_\d{2}-\d{2}-\d{2}).*$',
    r'^Set_Desglose_(Front|Back)_\d{2}-\d{2}-\d{2}.*$',
    r'PFAS\d{1}.*\.wiff$',
    r'Strata\d{2}.*\.wiff$',
    r'CLMS\d{3}.*\.wiff$',
    r'Acri_.*\.wiff$',
    r'^\d{6}_.*\.wiff$'
]
# Patrones específicos para archivos de líquidos (LC-MS)
datafile_liq =[
    r'PFAS\d{1}.*', # Ej: "PFAS1_240625.wiff"
    r'Strata\d{2}.*',
    r'CLMS\d{3}.*',
    r'Acri_.*',
    r'^\d{6}_.*'
]
# Mapeo de meses (para nombres de carpetas)
meses = {
    "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
    "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
    "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
}
# Diccionarios para control de estado y errores
error_anterior = {} # Registra últimos errores por equipo
alertas_mostradas = {"perdida_conexion": False} # Evita alertas duplicadas
text_estado_global = None # Referencia al widget de texto en la GUI

def buscar_carpeta(ruta_local, fecha_batch):
    """
    Busca carpetas de secuencias en una ruta local, retrocediendo en el tiempo si es necesario.
    
    Args:
        ruta_local (str): Ruta base donde buscar (ej: "\\servidor\Data\2024").
        fecha_batch (str): Fecha inicial en formato "DD-MM-AA".
    
    Returns:
        tuple: (carpetas_con_formato, idsiguiente, fecha_carpeta)
            - carpetas_con_formato: Lista de tuplas (nombre, timestamp, ruta) de carpetas válidas.
            - idsiguiente: Indica si hay secuencias posteriores (0: no, 1: sí).
            - fecha_carpeta: Fecha de la última carpeta encontrada.
    """
    # Lógica principal:
    # 1. Ajusta formato de fecha (ej: "25-06-2024" → "25-06-24").
    # 2. Busca en el mes actual y, si no hay resultados, en el mes anterior.
    # 3. Filtra carpetas que coincidan con los patrones `patterns`.
    """Ejemplo censurado: No implementa búsqueda real. Para más información contacta conmigo"""

    return [], 0, "01-01-2000"

# Definir las rutas por sección
def obtener_rutas(seccion, fecha, metodo, metodo_liquidos):
    """
    Devuelve las rutas y configuraciones por equipo según la sección (Semivol, Twister, etc.).
    
    Args:
        seccion (str): Área del laboratorio (ej: "Volátiles").
        fecha (str): Fecha de referencia para rutas anuales.
        metodo (str): Método analítico (ej: "CGM-020-a").
        metodo_liquidos (str): Método específico para LC-MS (opcional).
    
    Returns:
        dict: Estructura con rutas, tiempos de inyección, y archivos XML por equipo.
    """    
    año = datetime.now().year
    if metodo == "CGM-020-a": 
        tiempo4078 = 28
    elif metodo == "CGM-036-a": 
        tiempo4078 = 18
    elif metodo == "CGM-038-a": 
        tiempo4078 = 23
    if metodo_liquidos == "CLMS_030-a+CLMS_019-a": 
        tiempo_liq = 17
        metodo_liquidos = "CLMS_019+030-a"
    elif metodo_liquidos == "CLMS_028-a": 
        tiempo_liq = 19
    elif metodo_liquidos == "CLMS_002-a":
        tiempo_liq = 14
    elif metodo_liquidos == "CLMS_007-a":
        tiempo_liq = 14
    elif metodo_liquidos == "CLMS_011-a": 
        tiempo_liq = 4
    elif metodo_liquidos == "CLMS_023-a": 
        tiempo_liq = 10
    elif metodo_liquidos == "CLMS_029-a": 
        tiempo_liq = 18
    else:
        tiempo_liq = 0  

    if seccion == "Semivol":
        return {
            "rutas": {
                "EQ-001": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 30, 
                            "xml":r"\\ruta-red\d\MassHunter\GCMS\1\sequence\default.sequence.xml",
                            "local": rf"\\ruta-red\d\Data file\{año}"},
                "EQ-002": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 30, 
                            "xml":r"\\ruta-red\d\MassHunter\GCMS\1\sequence\SEMIVOL.sequence.xml",
                            "local": rf"\\ruta-red\d\DataFile\{año}"},
                "EQ-003": {"ruta": r"\\ruta-red\d (qqq-3)\MassHunter\GCMS\1", "tiempo_inyeccion": 30, 
                            "xml":r"\\ruta-red\d (qqq-3)\MassHunter\GCMS\1\sequence\Semivol.sequence.xml",
                            "local": rf"\\ruta-red\d (qqq-3)\Data file\{año}"}, 
                "EQ-004": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 30, 
                            "xml":r"\\ruta-red\d\MassHunter\GCMS\1\sequence\Semivolatiles.sequence.xml",
                            "local": rf"\\ruta-red\d\Data\{año}"}
            }
        }
    elif seccion == "Twister":
        return {
            "rutas": {
                "EQ-005": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 75,
                            "local": rf"\\ruta-red\d\Data File\Secuencias CGM-019-a\Secuencias {año}",
                            "xml2":r"\\ruta-red\d\MassHunter\GCMS\1\sequence\SEMIVOLATILES.SEQUENCE.xml",
                            "xml":r"\\ruta-red\C\GCMS\Msexe\masont_1.sequence.xml"},
                "EQ-006": {"ruta": r"\\ruta-red\D\MassHunter\GCMS\1", "tiempo_inyeccion": 75,
                            "local": rf"\\ruta-red\D\Secuencias CGM-019-a\Secuencias {año}",
                            "xml":r"\\ruta-red\D\MassHunter\GCMS\1\sequence\Semivolátiles.sequence.xml",},
                "EQ-007": {"ruta": r"\\ruta-red\D\MassHunter\GCMS\1", "tiempo_inyeccion": 50,
                            "local": rf"\\ruta-red\D\Data file\Secuencias CGM-031-a\Secuencias {año}",
                            "xml2":r"\\ruta-red\D\MassHunter\GCMS\1\sequence\TWISTER-CPTOS SEMIVOLATILES.sequence.xml",
                            "xml":r"\\ruta-red\C\GCMS\Msexe\masont_1.sequence.xml"},
                "EQ-008": {"ruta": r"\\ruta-red\d (EQ-008)\MassHunter\GCMS\1", "tiempo_inyeccion": 50,
                            "local": rf"\\ruta-red\d (EQ-008)\Data file\Secuencias CGM-031-a\Secuencias {año}",
                            "xml2":r"\\ruta-red\d (EQ-008)\MassHunter\GCMS\1\sequence\SEMIVOLATILES.SEQUENCE.xml",
                            "xml":r"\\ruta-red\C\GCMS\Msexe\masont_1.sequence.xml"},
                "EQ-009": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 69,
                            "local": rf"\\ruta-red\d\Secuencias {año}",
                            "xml2":r"\\ruta-red\d\MassHunter\GCMS\1\sequence\SEMIVOLATILES.SEQUENCE.xml",  
                            "xml":r"\\ruta-red\c\gcms\MSexe\masont_1.sequence.xml"},                            
                "EQ-010": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 50,
                            "local": rf"\\ruta-red\d\DATA FILE\Secuencias {año}",
                            "xml2":r"\\ruta-red\d\MassHunter\GCMS\1\sequence\Secuencia semivolatiles.SEQUENCE.xml",
                            "xml":r"\\ruta-red\c\GCMS\Msexe\masont_1.sequence.xml"},
                "EQ-011": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 21,
                            "local": rf"\\ruta-red\d\EQ-011\Data\CGM-011-a\{año}",
                            "local2": rf"\\ruta-red\d\EQ-011\Data\CGM-009-a\{año}",
                            "xml":r"\\ruta-red\d\EQ-011\Sequence\CGM-011-a.sequence.xml",
                            "xml2":r"\\ruta-red\d\EQ-011\Sequence\CGM-009-a.sequence.xml",
                            "xml3":r"\\ruta-red\d\EQ-011\Sequence\CGM-009-a y 022.sequence.xml"}
            }
        }
    elif seccion == "Volátiles":
        return {
            "rutas": {
                "EQ-012": {"ruta": r"\\ruta-red\D\MassHunter\GCMS\1", "tiempo_inyeccion": 24,
                            "excel":rf"\\ruta-red\D\Agilent 3194-3195\Data\{año}\HS",
                            "xml":r"\\ruta-red\c\GCMS\Msexe\masont_1.sequence.xml"},
                "EQ-013": {"ruta": r"\\ruta-red\D\MassHunter\GCMS\1", "tiempo_inyeccion": 28,
                            "excel":rf"\\ruta-red\cs-4101-4102\Data\{año}",
                            "xml":r"\\ruta-red\c\GCMS\Msexe\masont_1.sequence.xml"},
                "EQ-014": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 27,
                            "excel":rf"\\ruta-red\d\EQ-014-MS-FID\Data\HS-FID\{año}",
                            "excel2":rf"\\ruta-red\d\EQ-014-MS-FID\Data\HS-MS\{año}",
                            "xml":r"\\ruta-red\d\MassHunter\GCMS\1\sequence\HS_FID.sequence.xml",
                            "xml2":r"\\ruta-red\d\MassHunter\GCMS\1\sequence\HS_MS.sequence.xml"},
                "EQ-015": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 24,
                            "excel":rf"\\ruta-red\d\EQ-015-MS-FID\Data\FID\{año}",
                            "excel2":rf"\\ruta-red\d\EQ-015-MS-FID\Data\MS\{año}",
                            "xml":r"\\ruta-red\d\EQ-015-MS-FID\Secuencias\HS_FID.sequence.xml",
                            "xml2":r"\\ruta-red\d\EQ-015-MS-FID\Secuencias\HS_MS.sequence.xml"},
                "EQ-016": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 29,
                            "excel":rf"\\ruta-red\d\EQ-016\DATA\{año}",
                            "xml":r"\\ruta-red\d\EQ-016\sequence\CGM-026-a.sequence.xml"},
                "EQ-017": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 27,
                            "excel":rf"\\ruta-red\d\EQ-017\Data\{año}",
                            "xml":r"\\ruta-red\d\EQ-017\Secuencias\Secuencia-CGM-040-a.sequence.xml"},
                "EQ-018": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 23,
                            "excel":rf"\\ruta-red\d\EQ-018\Data\{año}",
                            "xml":r"\\ruta-red\d\EQ-018\Secuencias\CGM-040-a.sequence.xml",
                            "xml2":r"\\ruta-red\d\EQ-018\Secuencias\CGM-040-n.sequence.xml"},
                "EQ-019": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 23,
                            "excel":rf"\\ruta-red\d\EQ-019-SPME-HS\Data\{año}\HS",
                            "excel2":rf"\\ruta-red\d\EQ-019-SPME-HS\Data\{año}\SPME",
                            "xml":r"\\ruta-red\D\EQ-019-SPME-HS\Secuencias\Secuencia-CGM-040-a.sequence.xml",
                            "xml2":r"\\ruta-red\D\EQ-019-SPME-HS\Secuencias\Secuencia-CGM-038-a.sequence.xml"},
                "EQ-020": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 27,
                            "excel":rf"\\ruta-red\d\EQ-020-HS-MS\Data\{año}",
                            "xml":r"\\ruta-red\D\EQ-020-HS-MS\Secuencias\Secuencia-CGM-040-a.sequence.xml",
                            "xml2":r"\\ruta-red\D\EQ-020-HS-MS\Secuencias\Secuencia-CGM-040-n.sequence.xml"}
            }
        }
    elif seccion == "Fenoles":
        return {
            "rutas": {
                "EQ-021": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": tiempo4078,
                            "xml":r"\\ruta-red\c\GCMS\Msexe\masont_1.sequence.xml",
                            "local": rf"\\ruta-red\d\EQ-021-3195\Data\{año}\{metodo}"}, 
                "EQ-022": {"ruta": r"\\ruta-red\d\MassHunter\GCMS\1", "tiempo_inyeccion": 18,
                            "xml":r"\\ruta-red\d\EQ-022-CG-MS\Secuencias\CGM-039-a.sequence.xml",
                            "local": rf"\\ruta-red\d\EQ-022-CG-MS\DATA\FENOLES\{año}"}
            }
        }
    elif seccion == "Hidrocarburos":
        return {
            "rutas": {
                "EQ-023": {"ruta": rf"\\ruta-red\d\CDSProjects\LTM\Results\{año}", "tiempo_inyeccion": 8,
                            "log":r"\\ruta-red\c\ProgramData\Agilent\LogFiles\AcquisitionServer-16.log"}, 
                "EQ-024": {"ruta": rf"\\ruta-red\d\CDSProjects\LTM2\Results\{año}", "tiempo_inyeccion": 8,
                            "log":r"\\ruta-red\c\ProgramData\Agilent\LogFiles\AcquisitionServer-16.log"}
            }
        }
    elif seccion == "Líquidos":
        return {
            "rutas": {
                "EQ-025": {"ruta": rf"\\ruta-red\d\Analyst Data\Projects\{año}", "tiempo_inyeccion": tiempo_liq},
                "EQ-026": {"ruta": rf"\\ruta-red\D_EQ-026\Analyst Data\Projects\{año}", "tiempo_inyeccion": tiempo_liq},
                "EQ-027": {"ruta": rf"\\ruta-red\D_EQ-027\SCIEX OS Data\{metodo_liquidos} ({año})\Data", "tiempo_inyeccion": tiempo_liq},
                "EQ-028": {"ruta": rf"\\ruta-red\d\SCIEX OS Data\{metodo_liquidos} ({año})\Data", "tiempo_inyeccion": tiempo_liq},
                "EQ-029": {"ruta": rf"\\ruta-red\d\SCIEX OS Data\{metodo_liquidos} ({año})\Data", "tiempo_inyeccion": tiempo_liq}
            }
        }
    return {"rutas": {}}

# Función para leer el archivo mslogbk.HTM y buscar errores en los últimos 30 minutos
def buscar_errores(ruta, equipo, error_anterior, errores):
    """
    Busca errores en el archivo mslogbk.htm (log de MassHunter) de los últimos 20 minutos.
    
    Args:
        ruta (str): Ruta al archivo mslogbk.htm.
        equipo (str): Identificador del equipo (ej: "EQ-021").
        error_anterior (dict): Registro de errores previos para evitar duplicados.
        errores (list): Lista acumulativa de errores encontrados.
    
    Returns:
        list: Lista actualizada de errores (con formato "EQUIPO: Mensaje de error").
    """
    # Procesa el HTML con BeautifulSoup para extraer errores recientes.
    # Filtra por hora y evita repetir errores ya notificados.
    try:
        if not os.path.exists(ruta):
            return errores

        hora_buscar = datetime.now() - timedelta(minutes=20)
        patron_hora = re.compile(r"(\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2})")
        leer = False
        nueva_hora_error = None
        
        with open(ruta, 'r', encoding="utf-16") as file:
            for linea in file:
                match = patron_hora.search(linea)
                if match:
                    hora_linea = datetime.strptime(match.group(1), "%Y/%m/%d %H:%M:%S")
                    if hora_linea > hora_buscar:
                        leer = True
                    
                    if equipo in error_anterior and hora_linea <= error_anterior[equipo]:
                        leer = False  # Ignorar errores previos
                    else:
                        nueva_hora_error = hora_linea
                
                if leer and "error" in linea.lower():
                    mensaje_limpio = BeautifulSoup(linea, "html.parser").get_text()
                    errores.append(f"\n {equipo}: {mensaje_limpio.strip()}")
                    error_anterior[equipo] = nueva_hora_error  # Actualizar última hora de error detectado
    except Exception as e:
        print(f"Error al leer el archivo {equipo}: {e}")
    return errores

def obtener_id_volátiles(equipo):
    try:
        ruta_archivo = r"\\ruta-red\Datos\3. PRODUCCION\8. LABORATORIO CROMATOGRAFIA CS\23.GESTIÓN Y CONTROL\23.1.GESTIÓN Y CONTROL-SOIL-LAB\23.1.3.HOJAS DE CÁLCULO\Data files de equipos.xlsm"
        df = pd.read_excel(ruta_archivo, sheet_name=0, header=None)
        fila = df[df[0] == equipo]
        if not fila.empty:
            data_file = int(fila.iloc[0, 1])  # Obtener el valor de la segunda columna
            return data_file
        else: 
            return None
    except Exception as e:
        print(f"Error al leer el archivo: {e}")
        return None
    
def verificar_carpetas_creadas(carpetas_con_formato):
    carpetas_inyecciones = []
    saltadas = False  # Inicializar la variable de control
    ahora = datetime.now()  # Obtener la fecha y hora actual
    limite_tiempo = ahora - timedelta(minutes=20)  # Calcular el límite de tiempo (últimos 20 minutos)

    # Obtener las fechas de creación de las carpetas creadas en los últimos 20 minutos
    fechas_creacion = []
    for nombre_carpeta, timestamp, ruta_completa in carpetas_con_formato:
        fecha_creacion = datetime.fromtimestamp(timestamp)  # Convertir el timestamp a objeto datetime
        if fecha_creacion >= limite_tiempo:  # Filtrar carpetas creadas en los últimos 20 minutos
            if any(re.search(pattern, nombre_carpeta, re.IGNORECASE) for pattern in patterns):
                carpetas_inyecciones.append(nombre_carpeta)
                fechas_creacion.append(fecha_creacion)
    
    if len(fechas_creacion) < 2:
        return False

    # Ordenar las fechas de creación
    fechas_creacion.sort()

    # Comparar las diferencias de tiempo entre carpetas consecutivas que cumpla el patrón
    for i in range(1, len(fechas_creacion)):
        diferencia = (fechas_creacion[i] - fechas_creacion[i - 1]).total_seconds()  # Diferencia en segundos
        if diferencia < 120:  # Si la diferencia es menor a 2 minutos (120 segundos)
            saltadas = True
            break  # Salir del bucle si se encuentra una coincidencia
    return saltadas

def buscar_secuencias(data_file_names,fecha_batch,idsiguiente,fecha_inyeccion):
    # Buscar si hay dos secuencias programadas en cola
    secuencia_anterior = 0
    contador = 0
    for data_file_name in data_file_names[:-1]:  # Excluir el último, ya lo procesamos
        for pattern in patterns:
            match = re.search(pattern, data_file_name.text, re.IGNORECASE)
            if match:
                fecha_actual = match.group(1)
                secuencia_actual = int(match.group(2))
                if fecha_actual != fecha_batch and fecha_actual == fecha_inyeccion:
                    contador += 1
                            
            # Si la fecha es diferente a fecha_batch, actualizar secuencia_anterior
    if idsiguiente == 0:
        secuencia_anterior = contador
    return secuencia_anterior

def encontrar_carpeta_reciente(ruta_local):
    # Orden de prioridad (de más reciente a más antigua)
    carpetas_posibles = [
        "CLMS_028-a (Octubre-Diciembre)",
        "CLMS_028-a (Julio-Septiembre)",
        "CLMS_028-a (Abril-Junio)",
        "CLMS_028-a (Enero-Marzo)",
        "CLMS_007 (Octubre-Diciembre)",
        "CLMS_007 (Julio-Septiembre)",
        "CLMS_007 (Abril-Junio)",
        "CLMS_007 (Enero-Marzo)",
    ]
    
    for carpeta in carpetas_posibles:
        ruta_completa = Path(ruta_local) / carpeta / "Data"
        
        if not ruta_completa.exists():
            continue
            
        try:
            # Buscar TODOS los archivos válidos en la carpeta actual
            archivos_validos = [
                archivo for archivo in ruta_completa.iterdir() 
                if archivo.is_file() and any(re.search(pattern, archivo.name) for pattern in patterns)
            ]
            
            # Si hay archivos válidos, devolver el ÚLTIMO (o el más reciente por fecha)
            if archivos_validos:
                ultima_carpeta = max(archivos_validos, key=lambda f: f.stat().st_mtime)
                return ultima_carpeta  # Último archivo por orden alfabético
                # return max(archivos_validos, key=lambda f: f.stat().st_mtime)
                
        except (PermissionError, OSError) as e:
            print(f"Error accediendo {ruta_completa}: {e}")
            continue
    
    return None  # Si no se encontró ningún archivo válido

def leer_secuencia_liquidos(equipo):
    # Definir la ruta de búsqueda específica para el equipo
    ruta_equipo = Path(rf"\\ruta-red\Datos\3. PRODUCCION\8. LABORATORIO CROMATOGRAFIA CS\3. GAP LIQUIDO-MASAS\27. SECUENCIAS INYECCION\{equipo}")
    
    # Variables a devolver
    resultado = None
    last_data_file_name = None
    cola_equipo = {}
    
    # Verificar si la carpeta del equipo existe
    if not ruta_equipo.exists():
        print(f"No se encontró la carpeta para el equipo {equipo}")
        return None, None
    
    # Buscar archivos CSV o TXT con formato de fecha AAAA-MM-DD
    archivos_encontrados = []
    for ext in ['.csv', '.txt']:
        archivos_encontrados.extend(ruta_equipo.glob(f"????-??-??{ext}"))
    
    if not archivos_encontrados:
        print(f"No se encontraron archivos con formato de fecha en la carpeta del equipo {equipo}")
        return None, None
    
    # Ordenar archivos por fecha (más reciente primero)
    archivos_encontrados.sort(key=lambda x: x.stem, reverse=True)
    
    # Procesar el archivo más reciente para obtener resultado y last_data_file_name
    archivo_mas_reciente = archivos_encontrados[0]
    try:
        # Leer el archivo según su extensión
        if archivo_mas_reciente.suffix.lower() == '.csv':
            df = pd.read_csv(archivo_mas_reciente)
            columna = 'LC Method'
            columna_archivo = 'Data File'
        elif archivo_mas_reciente.suffix.lower() == '.txt':
            df = pd.read_csv(archivo_mas_reciente, sep='\t')
            columna = 'AcqMethod'
            columna_archivo = 'OutputFile'
        
        # Obtener last_data_file_name del archivo más reciente
        if columna_archivo in df.columns:
            valores_validos = df[columna_archivo].dropna()
            for valor in reversed(valores_validos):
                if any(re.match(patron, str(valor)) for patron in datafile_liq):
                    last_data_file_name = valor
                    break
        
        # Obtener resultado (LC Method/AcqMethod procesado)
        if columna in df.columns and not df.empty:
            primer_valor = df[columna].iloc[0]
            indice_a = primer_valor.lower().find('a')
            resultado = primer_valor[:indice_a + 1] if indice_a != -1 else None
    
    except Exception as e:
        print(f"Error al procesar el archivo más reciente {archivo_mas_reciente.name}: {e}")
    
    # Procesar hasta 4 archivos para el diccionario cola_equipo
    for archivo in archivos_encontrados[:4]:
        try:
            if archivo.suffix.lower() == '.csv':
                df_cola = pd.read_csv(archivo)
                col_archivo = 'Data File'
            elif archivo.suffix.lower() == '.txt':
                df_cola = pd.read_csv(archivo, sep='\t')
                col_archivo = 'OutputFile'
            
            current_last_data = None
            if col_archivo in df_cola.columns:
                valores_validos = df_cola[col_archivo].dropna()
                for valor in reversed(valores_validos):
                    if any(re.match(patron, str(valor)) for patron in datafile_liq):
                        current_last_data = valor
                        break
            
            cola_equipo[archivo.stem] = current_last_data
        
        except Exception as e:
            print(f"Error al procesar archivo {archivo.name} para cola_equipo: {e}")
            cola_equipo[archivo.stem] = None
    
    return resultado, cola_equipo

# Función para ejecutar la verificación de errores y actualizar la ventana de estado
estados_anteriores = {}
def ejecutar_verificacion(secciones, ventana_estado, text_estado):
    """
    Ejecuta la verificación periódica de equipos y actualiza la GUI.
    
    Args:
        secciones (list): Lista de secciones a monitorear (ej: ["Twister", "Líquidos"]).
        ventana_estado (tk.Toplevel): Ventana donde se muestra el estado.
        text_estado (tk.Text): Widget de texto para mostrar resultados.
    """
    # Flujo principal:
    # 1. Para cada sección y equipo:
    #   - Lee archivos XML/secuencias para obtener la muestra actual.
    #   - Verifica errores en logs (mslogbk.htm).
    #   - Calcula tiempo restante de la secuencia.
    # 2. Actualiza la GUI con colores:
    #   - Verde: Inyectando.
    #   - Rojo: Error.
    #   - Azul: Secuencia finalizada.
    # 3. Muestra alertas emergentes si hay cambios críticos.
    global estados_anteriores
    global id_ejecucion

    if id_ejecucion:
        ventana_estado.after_cancel(id_ejecucion)

    fecha_actual = datetime.now().strftime("%Y/%m/%d")
    todos_errores = []
    muestra_saltada = []
    equipo_parado = []
    perdida_conexion = []
    text_estado.config(state=tk.NORMAL)
    text_estado.delete(1.0, tk.END)
    hora_actual = datetime.now()
    hora_actual2 = hora_actual.strftime("%H:%M")
    hora_actualizacion = hora_actual + timedelta(minutes=15)
    hora_formateada = hora_actualizacion.strftime("%H:%M")
    text_estado.insert(tk.END, f"------- Actualizado a las: {hora_actual2}  |  Siguiente actualización: {hora_formateada} -------\n")

    for seccion in secciones:
        metodo = None
        if seccion == "Fenoles":
            ruta_fenoles = r"\\ruta-red\c\GCMS\Msexe\masont_1.sequence.xml"
            tree = ET.parse(ruta_fenoles)
            root = tree.getroot()
            namespaces = {'ns0': 'http://www.agilent.com/SequenceTable.xsd'}
            data_file_names = root.findall('.//ns0:Sequence/ns0:AcqMethodFileName', namespaces)
            metodo = data_file_names[-1].text.rstrip(".M")
            metodo = metodo[:9]

        datos_seccion = obtener_rutas(seccion, fecha_actual, metodo, None)

        # Insertar el título de la sección
        text_estado.insert(tk.END, f"\n=== {seccion} ===\n", "titulo")
        
        for equipo, info in datos_seccion["rutas"].items():
            parado = False
            conexion = False
            finalizado = False
            nombre_ultima_carpeta = None
            inyeccion_actual = None
            inyecciones_totales = None
            data_filename = None
            tiempo_total = None
            idactual = ""
            saltadas = None
            errores = []
            if equipo != "EQ-021" and seccion == "Fenoles": metodo = None

            hora_creacion = datetime.strptime("01-01-25", "%d-%m-%y")
            ruta = info["ruta"]  # Obtener la ruta del equipo
            tiempo_inyeccion = info["tiempo_inyeccion"]  # Obtener el tiempo de inyección

            #Lógica para Hidrocarburos
            try:
                if seccion == "Hidrocarburos":
                    ruta_log = info["log"]
                    ruta_local = info["ruta"]
                    inyecciones_cola = 0
                    ruta_completa = ruta_log
                    patron = r'\[ServerRunQueue\].*?UpdateInfo RQ Details: Label = (.*?): ItemValue = (.*)'
                    cola = r'\[SequenceRunValidator\]\[(.*?)\]\[Info\].*?\[ValidateAndCopyFiles\] Total injections for the sequence .*? is (\d+)'
                    try:
                        ruta_temporal = r"C:\temp\AcquisitionServer-16.log"
                        shutil.copy(ruta_log, ruta_temporal)
                        with open(ruta_temporal, "r", encoding="utf-8", errors="ignore") as file:
                            log_content = file.read()
                            
                            # Buscar todos los bloques que coincidan
                            matches = re.finditer(patron, log_content)
                            matches2 = re.finditer(cola, log_content)
                            
                            # Procesar todos los matches para obtener el último bloque completo
                            for match in matches:
                                label = match.group(1).strip()
                                value = match.group(2).strip()
                                
                                if label == "Current Injection":
                                    inyeccion_actual = int(value)
                                elif label == "Total Injections":
                                    inyecciones_totales = int(value)
                                elif label == "Start Time":
                                    inicio_secuencia = value                                    
                                    dt_inicio = parse(inicio_secuencia)
                                    dt_inicio = dt_inicio.replace(tzinfo=None)
                                elif label == "Data Filename":
                                    data_filename = value
                                    fecha_hidrocarb = re.search(r'^(\d{2}-\d{2}-\d{2})', data_filename)
                                    fecha_batch = fecha_hidrocarb.group(1)
                            for match2 in matches2:
                                hora_cola = match2.group(1).strip()
                                dt_inyecciones = datetime.fromisoformat(hora_cola)
                                dt_inicio = dt_inicio.replace(tzinfo=timezone(timedelta(hours=2)))
                                if dt_inyecciones > dt_inicio:
                                    inyecciones_cola += int(match2.group(2).strip())

                            carpetas_con_formato, idsiguiente, fecha_carpeta = buscar_carpeta(ruta_local, fecha_batch)
                            ultima_carpeta = max(carpetas_con_formato, key=lambda x: x[1])
                            nombre_ultima_carpeta = ultima_carpeta[0]
                            timestamp_creacion = ultima_carpeta[1]
                            hora_creacion = datetime.fromtimestamp(timestamp_creacion)  
                            t_iny = timedelta(minutes=tiempo_inyeccion) - timedelta(minutes=1)
                            t_restante = t_iny - (datetime.now() - hora_creacion)
                            tiempo_total = (inyecciones_totales - inyeccion_actual + inyecciones_cola) * t_iny + t_restante
                            dias = tiempo_total.days
                            horas = tiempo_total.seconds // 3600 + dias*24
                            minutos = (tiempo_total.seconds % 3600) // 60
                    except:
                        print("Traceback completo:")
                        traceback.print_exc()
                        continue   
                
                #Lógica para Líquidos
                elif seccion == "Líquidos":
                    metodo_liquidos, lista_secuencia = leer_secuencia_liquidos(equipo)
                    if metodo_liquidos == "CLMS_030-a":
                        metodo_liquidos = "CLMS_030-a+CLMS_019-a"
                    elif metodo_liquidos == "CLMS 007-a":
                        metodo_liquidos = "CLMS_007-a" 
                    elif metodo_liquidos == "CLMS_002-A":
                        metodo_liquidos = "CLMS_002-a"  

                    secuencias_liq = {}
                    last_data_file_name = None
                    max_fecha = None
                    for fecha_str, nombre_archivo in lista_secuencia.items():
                        coincidencia = re.search(r'(?:_)(\d+)$', str(nombre_archivo))
                        if coincidencia:
                            secuencias_liq[fecha_str] = int(coincidencia.group(1))
                        else:
                            secuencias_liq[fecha_str] = None
                        fecha_dt = datetime.strptime(fecha_str, '%Y-%m-%d')
                        if max_fecha is None or fecha_dt > max_fecha:
                            max_fecha = fecha_dt
                            last_data_file_name = nombre_archivo

                    datos_seccion = obtener_rutas(seccion, fecha_actual, metodo, metodo_liquidos)
                    tiempo_inyeccion = datos_seccion["rutas"][equipo]["tiempo_inyeccion"]
                    for equipo2, info2 in datos_seccion["rutas"].items():
                        if equipo2 == equipo:
                            ruta_completa = info2["ruta"]
                    if equipo == "EQ-025" or equipo == "EQ-026":
                        ruta_local = info["ruta"]
                        ruta_completa = ruta_local
                        ultima_carpeta = encontrar_carpeta_reciente(ruta_local)
                        timestamp_creacion = os.path.getctime(ultima_carpeta)
                        hora_creacion = datetime.fromtimestamp(timestamp_creacion)
                        nombre_ultima_carpeta = os.path.splitext(os.path.basename(ultima_carpeta))[0]
                        fecha_inyeccion_str = nombre_ultima_carpeta[:6]
                        fecha_liquidos = datetime.strptime(fecha_inyeccion_str, '%y%m%d').strftime('%Y-%m-%d')
                        ultima_sample = 0
                        for fecha_str in secuencias_liq:
                            fecha_dict = datetime.strptime(fecha_str, '%Y-%m-%d')
                            fecha_inyeccion_dt = datetime.strptime(fecha_liquidos, '%Y-%m-%d')
                            if fecha_dict >= fecha_inyeccion_dt:
                                ultima_sample += secuencias_liq[fecha_str]  # Mantienes tu lógica original

                        timestamp_creacion = os.path.getctime(ultima_carpeta)
                    else:
                        archivos_encontrados = []
                        with os.scandir(ruta_completa) as entries:
                            for entry in entries:
                                if entry.is_file() and any(re.search(patron, entry.name) for patron in datafile_liq):
                                    archivos_encontrados.append(entry)
                            if archivos_encontrados:
                                mas_reciente = max(archivos_encontrados, key=lambda f: f.stat().st_mtime)
                                nombre_ultima_carpeta = Path(mas_reciente).name.split('.')[0]
                                fecha_inyeccion_str = nombre_ultima_carpeta[:6]
                                fecha_liquidos = datetime.strptime(fecha_inyeccion_str, '%y%m%d').strftime('%Y-%m-%d')
                                ultima_sample = 0
                                for fecha_str in secuencias_liq:
                                    fecha_dict = datetime.strptime(fecha_str, '%Y-%m-%d')
                                    fecha_inyeccion_dt = datetime.strptime(fecha_liquidos, '%Y-%m-%d')
                                    if fecha_dict >= fecha_inyeccion_dt:
                                        ultima_sample += secuencias_liq[fecha_str]  # Mantienes tu lógica original
                                timestamp_creacion = os.path.getctime(mas_reciente)
                        hora_creacion = datetime.fromtimestamp(timestamp_creacion)
                    coincidencia = re.search(r'(?:_)(\d+)$', nombre_ultima_carpeta)
                    ultima_inyeccion = int(coincidencia.group(1))
                    t_iny = timedelta(minutes=tiempo_inyeccion) - timedelta(minutes=1)
                    t_restante = t_iny - (datetime.now() - hora_creacion)
                    tiempo_total = (ultima_sample - ultima_inyeccion) * t_iny + t_restante
                    dias = tiempo_total.days
                    horas = tiempo_total.seconds // 3600 + dias*24
                    minutos = (tiempo_total.seconds % 3600) // 60

                #Lógica para MassHunter
                else:
                    ruta_completa = os.path.join(ruta, "mslogbk.htm")
                    
                    errores = buscar_errores(ruta_completa, equipo, error_anterior, errores)

                    if "xml" in info:
                        try:
                            try:
                                ruta_xml = info["xml"]
                                if "local" in info:
                                    ruta_local = info["local"]
                                if "excel" in info:
                                    ruta_local = info["excel"]
                                tiempo_inyeccion = info["tiempo_inyeccion"]

                                # Parsear el XML
                                tree = ET.parse(ruta_xml)
                                root = tree.getroot()
                                namespaces = {'ns0': 'http://www.agilent.com/SequenceTable.xsd'}
                                data_file_names = root.findall('.//ns0:Sequence/ns0:DataFileName', namespaces)
                                last_data_file_name = data_file_names[-1].text
                            except:
                                last_data_file_name = None
                            
                            # Extraer la fecha del nombre del archivo
                            if seccion == "Twister" and equipo != "EQ-011":
                                fecha_batch = "01-01-2025"
                            else:
                                fecha_batch = "01-01-25"
                            data_file = None

                            if last_data_file_name is not None:
                                for pattern in patterns:
                                    match = re.search(pattern, last_data_file_name, re.IGNORECASE)
                                    if match:
                                        fecha_batch = match.group(1)
                                        data_file = int(match.group(2))
                                        break  # Termina el bucle después del primer match

                            # Asegurar que data_file no sea None antes de usarlo
                            if data_file is None:
                                data_file = 0
                            #Si es un equipo con dos posibles xml, buscar el que tenga la secuencia más reciente
                            if "xml2" in info:
                                try:
                                    ruta_xml2 = info["xml2"]
                                    tree2 = ET.parse(ruta_xml2)
                                    root2 = tree2.getroot()
                                    namespaces = {'ns0': 'http://www.agilent.com/SequenceTable.xsd'}
                                    data_file_names = None
                                    fecha_batch_2 = None
                                    data_file_2 = None
                                    data_file_names = root2.findall('.//ns0:Sequence/ns0:DataFileName', namespaces)
                                    last_data_file_name_2 = data_file_names[-1].text
                                    for pattern in patterns:
                                        match = re.search(pattern, last_data_file_name_2, re.IGNORECASE)
                                        if match:
                                            fecha_batch_2 = match.group(1)
                                            data_file_2 = int(match.group(2))
                                            break  # Termina el bucle después del primer match
                                    try:
                                        fecha_2 = datetime.strptime(fecha_batch_2, "%d-%m-%Y")
                                    except:
                                        fecha_2 = datetime.strptime(fecha_batch_2, "%d-%m-%y")  
                                    try:
                                        fecha_1 = datetime.strptime(fecha_batch, "%d-%m-%Y")
                                    except:
                                        fecha_1 = datetime.strptime(fecha_batch, "%d-%m-%y")                            
                                    if fecha_2 - fecha_1 > timedelta(0) or last_data_file_name is None:
                                        fecha_batch = fecha_batch_2
                                        data_file = data_file_2
                                        last_data_file_name = last_data_file_name_2              
                                        if seccion == "Volátiles" and "excel2" in info: ruta_local = info["excel2"]
                                        if "local2" in info:
                                            ruta_local = info["local2"]
                                        tree = tree2
                                        root = root2   
                                except:
                                    print("Traceback completo:")
                                    traceback.print_exc()
                                    continue    
                            if "xml3" in info:
                                try:
                                    ruta_xml3 = info["xml3"]
                                    tree3 = ET.parse(ruta_xml3)
                                    root3 = tree3.getroot()
                                    namespaces = {'ns0': 'http://www.agilent.com/SequenceTable.xsd'}
                                    data_file_names = None
                                    fecha_batch_3 = None
                                    data_file_3 = None
                                    data_file_names = root3.findall('.//ns0:Sequence/ns0:DataFileName', namespaces)
                                    last_data_file_name_3 = data_file_names[-1].text
                                    for pattern in patterns:
                                        match = re.search(pattern, last_data_file_name_3, re.IGNORECASE)
                                        if match:
                                            fecha_batch_3 = match.group(1)
                                            data_file_3 = int(match.group(2))
                                            break  # Termina el bucle después del primer match
                                    try:
                                        fecha_3 = datetime.strptime(fecha_batch_3, "%d-%m-%Y")
                                    except:
                                        fecha_3 = datetime.strptime(fecha_batch_3, "%d-%m-%y")  
                                    try:
                                        fecha_1 = datetime.strptime(fecha_batch, "%d-%m-%Y")
                                    except:
                                        fecha_1 = datetime.strptime(fecha_batch, "%d-%m-%y")                            
                                    if fecha_3 - fecha_1 > timedelta(0) or last_data_file_name is None:
                                        fecha_batch = fecha_batch_3
                                        data_file = data_file_3
                                        last_data_file_name = last_data_file_name_3
                                        if "local3" in info:
                                            ruta_local = info["local3"]
                                        tree = tree3
                                        root = root3   
                                except:
                                    print("Traceback completo:")
                                    traceback.print_exc()
                                    continue  

                            # Llamar a la función buscar_carpeta
                            carpetas_con_formato, idsiguiente, fecha_carpeta = buscar_carpeta(ruta_local, fecha_batch)

                            if seccion == "Volátiles" or metodo == "CGM-038-a" or metodo == "CGM-020-a":
                                data_file = obtener_id_volátiles(equipo)

                            if carpetas_con_formato:
                                ultima_carpeta = max(carpetas_con_formato, key=lambda x: x[1])
                                nombre_ultima_carpeta = ultima_carpeta[0]
                                limpiezas = 0
                                #Contar cuantas limpiezas hay en métodos 42-43 para hacer cálculo de tiempo
                                if seccion == "Semivol":
                                    sequences = root.findall('.//ns0:Sequence', namespaces)
                                    encontrado = False
                                    for seq in sequences:
                                        data_file_namespace = seq.find('ns0:DataFileName', namespaces)
                                        acq_method = seq.find('ns0:AcqMethodFileName', namespaces)
                                        # Si encontramos el DataFileName que coincide con nombre_ultima_carpeta
                                        if data_file_namespace is not None and nombre_ultima_carpeta in data_file_namespace.text:
                                            encontrado = True
                                        
                                        # Si ya encontramos el DataFileName objetivo, contamos las limpiezas
                                        if encontrado and acq_method is not None and acq_method.text == "Limpieza.M":
                                            limpiezas += 1

                                timestamp_creacion = ultima_carpeta[1]
                                hora_creacion = datetime.fromtimestamp(timestamp_creacion)
                                t_iny = timedelta(minutes=tiempo_inyeccion) - timedelta(minutes=2)
                                t_restante = t_iny - (datetime.now() - hora_creacion)

                                if seccion == "Twister":
                                    # Obtener todos los elementos <DataFileName>
                                    data_files = root.findall(".//ns0:DataFileName", namespaces)
                                    
                                    # Extraer los valores de texto de los elementos <DataFileName>
                                    valores_data_files = [elem.text for elem in data_files]
                                    
                                    # Encontrar la posición de la última carpeta procesada
                                    pos = next((i for i, valor in enumerate(valores_data_files) if nombre_ultima_carpeta.startswith(valor)), None)
                                    
                                    # Calcular el número de archivos restantes
                                    if pos is not None and pos < len(data_files) - 1:
                                        count_remaining = len(data_files) - (pos + 1)
                                    else:
                                        count_remaining = 0
                                    
                                    # Calcular el tiempo total restante
                                    tiempo_total = count_remaining * t_iny + t_restante
                                    
                                else:
                                    for pattern in patterns:
                                        match2 = re.search(pattern, nombre_ultima_carpeta, re.IGNORECASE)
                                        if not idactual:
                                            fecha_inyeccion = match2.group(1)
                                            idactual = int(match2.group(2))
                                            break  # Termina el bucle después del primer match
                                    if fecha_carpeta != 0 and fecha_inyeccion != fecha_carpeta: idsiguiente = 1
                                    secuencia_anterior = buscar_secuencias(data_file_names,fecha_batch,idsiguiente,fecha_inyeccion)
                                    tiempo_total = (data_file - idactual + secuencia_anterior) * t_iny + t_restante - timedelta(minutes=limpiezas*10)
                                dias = tiempo_total.days
                                horas = tiempo_total.seconds // 3600 + dias*24
                                minutos = (tiempo_total.seconds % 3600) // 60
                                if seccion == "Volátiles":
                                    saltadas = verificar_carpetas_creadas(carpetas_con_formato)
                            if not carpetas_con_formato:
                                parado = True
                        except Exception as e:
                            print(f"Error: {e}")
                            text_estado.insert(tk.END, f"\n{equipo}: Fallo en el XML o nombre de Sample", "rojo")
                            print("Traceback completo:")
                            traceback.print_exc()
                            continue

                    if seccion == "Twister":
                        try:
                            ruta_local = info["local"]
                            tiempo_inyeccion = info["tiempo_inyeccion"]
                            fecha_batch = datetime.now().strftime("%d-%m-%y")  
                            carpetas_con_formato, idsiguiente, fecha_carpeta = buscar_carpeta(ruta_local, fecha_batch)
                            if carpetas_con_formato:
                                ultima_carpeta = max(carpetas_con_formato, key=lambda x: x[1])
                                nombre_ultima_carpeta = ultima_carpeta[0]
                                timestamp_creacion = ultima_carpeta[1]
                                hora_creacion = datetime.fromtimestamp(timestamp_creacion) 
                            else:
                                parado = True
                        except Exception as e:
                            print(f"Error: {e}")
                            print("Traceback completo:")
                            traceback.print_exc()   
                    
                #Verificar si el equipo está parado o no
                if seccion == "Semivol" or seccion == "Twister" or seccion == "Volátiles" or  seccion == "Fenoles" or seccion == "Hidrocarburos" or seccion == "Líquidos":
                    if datetime.now() > hora_creacion + timedelta(minutes=tiempo_inyeccion):
                        if seccion != "Hidrocarburos" and seccion != "Líquidos" and len(errores) != 0:
                            continue
                        if nombre_ultima_carpeta and seccion != "Hidrocarburos":
                            if not nombre_ultima_carpeta.startswith(last_data_file_name) and not parado:
                                conexion = True
                                emergente_conexion = True
                            else:
                                parado = True
                        if nombre_ultima_carpeta and seccion == "Hidrocarburos":
                            if inyecciones_totales - inyeccion_actual > 1:
                                conexion = True
                                emergente_conexion = True
                            else:
                                parado = True
                                emergente_conexion = False
            except Exception as e:
                text_estado.insert(tk.END, f"Error general al procesar el equipo {equipo}: {e}", "rojo")
                print("Traceback completo:")
                traceback.print_exc()            
            if equipo in estados_anteriores:
                estado_anterior = estados_anteriores[equipo]
            else:
                estado_anterior = "Desconocido"  # O "Equipo parado" según lo que tenga más sentido
                if not parado and not errores and not conexion:
                    estados_anteriores[equipo] = "Inyectando"
                elif conexion:
                    estados_anteriores[equipo] = "Secuencia incompleta"
                else:
                    estados_anteriores[equipo] = "Equipo parado"
            if estado_anterior == "Inyectando" and parado == True:
                finalizado = True
            # Lógica para pérdida de conexión (solo alerta la primera vez)
            if estado_anterior != "Secuencia incompleta" and conexion:
                # ¡Nueva pérdida de conexión! Mostrar alerta
                emergente_conexion = True
                alertas_mostradas["perdida_conexion"] = True  # Marcar como mostrada
            else:
                # Se recuperó la conexión, resetear el flag para futuras alertas
                alertas_mostradas["perdida_conexion"] = False
                emergente_conexion = False

            # Asegurar que siempre se actualiza el estado
            if not parado and not errores and not conexion:
                estados_anteriores[equipo] = "Inyectando"
            elif conexion:
                estados_anteriores[equipo] = "Secuencia incompleta"
            else:
                estados_anteriores[equipo] = "Equipo parado"

            # Mostrar el resultado en el widget de texto
            if errores:
                todos_errores.extend(errores)
                text_estado.insert(tk.END, f"\n{equipo}: Error detectado", "rojo")
                if tiempo_total is not None and int(tiempo_total.total_seconds()) > 0:
                    text_estado.insert(tk.END, f" (Tiempo restante: {horas} horas y {minutos} minutos)")
                elif (seccion == "Semivol" or seccion == "Volátiles" or seccion == "Fenoles" or seccion == "Twister") and not parado:
                    text_estado.insert(tk.END, f" No has guardado la sample")
            elif parado:
                text_estado.insert(tk.END, f"\n{equipo}: Secuencia finalizada", "azul")
            elif conexion:
                text_estado.insert(tk.END, f"\n{equipo}: Secuencia incompleta", "naranja")
            elif os.path.exists(ruta_completa):
                text_estado.insert(tk.END, f"\n{equipo}: Inyectando", "verde")
                if tiempo_total is not None and int(tiempo_total.total_seconds()) > 0:
                    text_estado.insert(tk.END, f" (Tiempo restante: {horas} horas y {minutos} minutos)")
                elif (seccion == "Semivol" or seccion == "Volátiles" or seccion == "Fenoles" or seccion == "Twister") and not parado:
                    text_estado.insert(tk.END, f" No has guardado la sample")
            else:
                text_estado.insert(tk.END, f"\n{equipo}: No se ha podido acceder a la ruta", "rojo")
            if finalizado:
                equipo_parado.append(f"\n {equipo}: Secuencia finalizada")
            if conexion:
                perdida_conexion.append(f"\n {equipo}: Secuencia incompleta")
            if saltadas:
                muestra_saltada.append(f"\n {equipo}: Muestras saltadas")
                text_estado.insert(tk.END, f" Muestras saltadas", "rojo")


        
        # Insertar un salto de línea entre secciones
        text_estado.insert(tk.END, "\n")
    
    text_estado.config(state=tk.DISABLED)
    if todos_errores or equipo_parado or muestra_saltada or (emergente_conexion and alertas_mostradas["perdida_conexion"]):
        if emergente_conexion:
            alertas_mostradas["perdida_conexion"] = False
        ventana_alerta = tk.Toplevel()
        ventana_alerta.title("¡Alerta!")
        ventana_alerta.geometry("400x300")
        ventana_alerta.attributes("-topmost", True)  # Mantiene la ventana en primer plano

        label_mensaje = tk.Label(ventana_alerta, text="Se detectaron los siguientes cambios:", font=("Arial", 12, "bold"))
        label_mensaje.pack(pady=10)

        text_errores = tk.Text(ventana_alerta, height=10, width=50, wrap=tk.WORD)
        if todos_errores:
            text_errores.insert(tk.END, "\n".join(todos_errores))
        if equipo_parado:
            text_errores.insert(tk.END, "\n".join(equipo_parado))
        if emergente_conexion:
            text_errores.insert(tk.END, "\n".join(perdida_conexion))
        if muestra_saltada:
            text_errores.insert(tk.END, "\n".join(muestra_saltada))
        text_errores.config(state=tk.DISABLED)  # Para que no se pueda editar
        text_errores.pack(padx=10, pady=10)

        btn_cerrar = tk.Button(ventana_alerta, text="Cerrar", command=ventana_alerta.destroy)
        btn_cerrar.pack(pady=10)
    
    id_ejecucion = ventana_estado.after(900000, ejecutar_verificacion, secciones, ventana_estado, text_estado)

# Crear la interfaz gráfica
def crear_interfaz():
    """
    Crea la ventana principal de la aplicación con:
    - Checkboxes para seleccionar secciones.
    - Botones para iniciar/exportar/salir.
    - Ventana secundaria para mostrar el estado.
    """
    # Configura ventana con tkinter/ttk:
    ventana = tk.Tk()
    ventana.title("Monitor de Errores")
    # ... (widgets y estilos)
    ventana = tk.Tk()
    ventana.title("Monitor de Errores")

    ventana.configure(padx=40, pady=25)

    global id_ejecucion
    id_ejecucion = None

    frame_centro = ttk.Frame(ventana)
    frame_centro.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

    selection_frame = ttk.LabelFrame(frame_centro, text="Seleccionar Secciones a Monitorizar")
    selection_frame.pack(fill=tk.X, pady=5)

    secciones = ["Semivol", "Twister", "Volátiles", "Fenoles", "Hidrocarburos", "Líquidos"]
    vars_secciones = {seccion: tk.BooleanVar() for seccion in secciones}

    for seccion in secciones:
        chk = ttk.Checkbutton(selection_frame, text=seccion, variable=vars_secciones[seccion])
        chk.pack(anchor=tk.W, padx=5, pady=2)

    def iniciar_monitoreo():
        global text_estado_global
        "Función censurada, para saber más, contacta conmigo."
        ventana = tk.Tk()
        ventana.title("Monitor (versión censurada)")
        tk.Label(ventana, text="Ejemplo sin lógica real.").pack()
        ventana.mainloop()

def exportar_a_excel():
    """
    Exporta el estado actual de los equipos a un archivo Excel.
    - Usa openpyxl para modificar el archivo sin perder macros.
    - Actualiza columnas como "Estado" y "Hora Fin".
    """
    try:
        if text_estado_global is None:
            messagebox.showerror("Error", "No hay datos para exportar. Primero inicia el monitoreo.")
            return
            
        contenido = text_estado_global.get("1.0", tk.END).splitlines()
        
        equipos = []
        for linea in contenido:
            if ":" in linea:
                partes = linea.split(":", 1)
                if len(partes) == 2:
                    equipo = partes[0].strip()
                    estado = partes[1].strip()
                    
                    # Determinar el estado base
                    if "Secuencia finalizada" in estado:
                        estado_base = "Parado"
                    elif "Secuencia incompleta" in estado:
                        estado_base = "Parado"
                    elif "Inyectando" in estado:
                        estado_base = "Inyectando"
                    else:
                        continue  # Saltar líneas no relevantes
                    
                    # Extraer tiempo restante si existe
                    hora_fin = "N/A"
                    if estado_base == "Inyectando":
                        match = re.search(r"(\d+) horas y (\d+) minutos", estado)
                        if match:
                            horas = int(match.group(1))
                            minutos = int(match.group(2))
                            hora_fin = (datetime.now() + timedelta(hours=horas, minutes=minutos)).strftime("%A %H:%M")
                    
                    equipos.append({
                        "Equipo": equipo,
                        "Estado": estado_base,
                        "Hora Fin (estimada)": hora_fin
                    })
        
        if not equipos:
            messagebox.showwarning("Advertencia", "No se encontraron datos válidos para exportar")
            return
            
        # Ruta del archivo Excel existente
        ruta_excel = r"\\ruta-red\Datos\3. PRODUCCION\8. LABORATORIO CROMATOGRAFIA CS\0. GENERAL\CONTROL EQUIPOS SÁBADOS.xlsm"
        
        # Cargar el libro de trabajo existente
        wb = load_workbook(ruta_excel, keep_vba=True)
        ws = wb.active  # Asume que trabajamos con la primera hoja
        
        # Obtener el día actual en formato dd-mm
        dia_actual = datetime.now().strftime("%d-%m")
        
        # Buscar coincidencias en la columna S (columna 19 en openpyxl)
        for row in ws.iter_rows(min_row=2):  # Asumiendo que la fila 1 es el encabezado
            equipo_excel = row[18].value  # Columna S (índice 18 ya que empieza en 0)
            
            if equipo_excel:
                # Buscar si este equipo está en nuestros datos
                for dato in equipos:
                    if str(equipo_excel).strip() == str(dato["Equipo"]).strip():
                        # Actualizar columna U (índice 20)
                        if dato["Estado"] == "Inyectando":
                            row[20].value = f"{dato['Estado']}. Finalización {dato['Hora Fin (estimada)']}"
                        else:
                            row[20].value = dato['Estado']
                        
                        # Actualizar columna AB (índice 27)
                        row[27].value = dia_actual
                        break
        
        # Guardar los cambios
        wb.save(ruta_excel)
        
        messagebox.showinfo("Éxito", f"Datos actualizados en:\n{ruta_excel}")
        
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar: {str(e)}")
        print("Error completo:", traceback.format_exc())

if __name__ == "__main__":
    crear_interfaz()
